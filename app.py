import os
import sys
import warnings
import streamlit as st
import tempfile
import pandas as pd
import lancedb
import traceback
import logging
import requests
import re
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Suppress warnings
warnings.filterwarnings("ignore")

# Set up minimal logging for startup
logging.basicConfig(level=logging.INFO, 
                   format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Configure the Streamlit page FIRST (before any heavy operations)
st.set_page_config(
    page_title="Document Processor", 
    layout="wide",
    initial_sidebar_state="expanded"
)

# Page header
st.title("Document Processor")
st.markdown("Process, search, and chat with your documents using OpenAI or Hugging Face models")

# Quick loading message
with st.spinner("Initializing application..."):
    # Set up environment variables with minimal overhead
    cache_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "cache"))
    os.makedirs(cache_dir, exist_ok=True)
    os.environ["HF_HOME"] = cache_dir
    os.environ["OMP_NUM_THREADS"] = "4"
    os.environ["TOKENIZERS_PARALLELISM"] = "false"

# Initialize session state for API keys and model preferences - Load from .env by default
if "openai_api_key" not in st.session_state:
    st.session_state.openai_api_key = os.getenv("OPENAI_API_KEY", "")
    
if "hf_api_key" not in st.session_state:
    st.session_state.hf_api_key = os.getenv("HF_API_KEY", "")
    
if "api_provider" not in st.session_state:
    st.session_state.api_provider = "Hugging Face"
    
if "embedding_model" not in st.session_state:
    st.session_state.embedding_model = "sentence-transformers/all-MiniLM-L6-v2"
    
if "llm_model" not in st.session_state:
    st.session_state.llm_model = "microsoft/DialoGPT-medium"

if "searched_models" not in st.session_state:
    st.session_state.searched_models = []

if "favorite_models" not in st.session_state:
    st.session_state.favorite_models = []

if "modules_loaded" not in st.session_state:
    st.session_state.modules_loaded = False

# LAZY LOADING FUNCTIONS - Only import when needed
@st.cache_resource
def load_embedding_functions():
    """Lazy load embedding functions only when needed"""
    try:
        # Add the knowledge/docling folder to the Python path
        sys.path.append(os.path.join(os.path.dirname(__file__), "knowledge", "docling"))
        
        from knowledge.docling.embedding import (
            process_document, 
            answer_with_rag, 
            get_available_embedding_models,
            get_available_llm_models,
            OpenAIEmbedder
        )
        return {
            'process_document': process_document,
            'answer_with_rag': answer_with_rag,
            'get_available_embedding_models': get_available_embedding_models,
            'get_available_llm_models': get_available_llm_models,
            'OpenAIEmbedder': OpenAIEmbedder
        }
    except ImportError as e:
        st.error(f"Import error: {str(e)}")
        return None

@st.cache_resource
def load_huggingface_client():
    """Lazy load Hugging Face client"""
    try:
        from huggingface_hub import InferenceClient
        return InferenceClient
    except ImportError:
        st.error("Hugging Face Hub not installed")
        return None

@st.cache_resource
def load_openai_client():
    """Lazy load OpenAI client"""
    try:
        from openai import OpenAI
        return OpenAI
    except ImportError:
        st.error("OpenAI package not installed")
        return None

# Enhanced Excel file processing function
def process_excel_file(file_path):
    """Process Excel files (.xlsx and .xls) and convert to text"""
    try:
        file_ext = os.path.splitext(file_path)[1].lower()
        text_content = []
        
        if file_ext == '.xlsx':
            try:
                # Try openpyxl first (for .xlsx files)
                import openpyxl
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    text_content.append(f"\n=== Sheet: {sheet_name} ===\n")
                    
                    # Process rows and columns
                    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                        row_text = []
                        for cell_idx, cell in enumerate(row, 1):
                            if cell is not None and str(cell).strip():
                                # Include column letter for better context
                                col_letter = openpyxl.utils.get_column_letter(cell_idx)
                                row_text.append(f"{col_letter}{row_idx}: {str(cell).strip()}")
                        
                        if row_text:
                            text_content.append(" | ".join(row_text))
                    
                    text_content.append("\n")  # Add space between sheets
                    
            except ImportError:
                # Fallback to pandas if openpyxl is not available
                try:
                    excel_file = pd.ExcelFile(file_path)
                    for sheet_name in excel_file.sheet_names:
                        df = pd.read_excel(file_path, sheet_name=sheet_name)
                        text_content.append(f"\n=== Sheet: {sheet_name} ===\n")
                        
                        # Convert DataFrame to text
                        for idx, row in df.iterrows():
                            row_text = []
                            for col_name, value in row.items():
                                if pd.notna(value) and str(value).strip():
                                    row_text.append(f"{col_name}: {str(value).strip()}")
                            
                            if row_text:
                                text_content.append(" | ".join(row_text))
                        
                        text_content.append("\n")
                        
                except Exception as pandas_error:
                    return f"Error processing .xlsx file: {str(pandas_error)}\nPlease install openpyxl: pip install openpyxl"
                            
        elif file_ext == '.xls':
            try:
                # For .xls files, use pandas or xlrd
                excel_file = pd.ExcelFile(file_path)
                for sheet_name in excel_file.sheet_names:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    text_content.append(f"\n=== Sheet: {sheet_name} ===\n")
                    
                    # Convert DataFrame to text
                    for idx, row in df.iterrows():
                        row_text = []
                        for col_name, value in row.items():
                            if pd.notna(value) and str(value).strip():
                                row_text.append(f"{col_name}: {str(value).strip()}")
                        
                        if row_text:
                            text_content.append(" | ".join(row_text))
                    
                    text_content.append("\n")
                    
            except Exception as e:
                return f"Error processing .xls file: {str(e)}\nPlease install xlrd: pip install xlrd"
        
        return "\n".join(text_content)
        
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        return f"Error processing Excel file: {str(e)}"

# Function to search Hugging Face models
@st.cache_data(ttl=3600)  # Cache for 1 hour
def search_huggingface_models(query, task_filter=None, library_filter=None, limit=20):
    """Search for models on Hugging Face Hub"""
    try:
        url = "https://huggingface.co/api/models"
        params = {
            "search": query,
            "limit": limit,
            "sort": "downloads",
            "direction": -1
        }
        
        if task_filter:
            params["filter"] = task_filter
        if library_filter:
            params["library"] = library_filter
            
        response = requests.get(url, params=params, timeout=10)
        
        if response.status_code == 200:
            models = response.json()
            
            # Filter for text generation and conversational models
            text_models = []
            for model in models:
                model_id = model.get("id", "")
                tags = model.get("tags", [])
                pipeline_tag = model.get("pipeline_tag", "")
                
                # Check if it's a text generation or conversational model
                if (pipeline_tag in ["text-generation", "conversational", "text2text-generation"] or
                    any(tag in ["text-generation", "conversational", "chat"] for tag in tags) or
                    any(keyword in model_id.lower() for keyword in ["gpt", "dialog", "chat", "conversation", "llama", "mistral", "flan"])):
                    
                    text_models.append({
                        "id": model_id,
                        "downloads": model.get("downloads", 0),
                        "likes": model.get("likes", 0),
                        "pipeline_tag": pipeline_tag,
                        "tags": tags
                    })
            
            # Sort by downloads
            text_models.sort(key=lambda x: x["downloads"], reverse=True)
            return text_models[:limit]
        else:
            return []
    except Exception as e:
        logger.error(f"Error searching models: {str(e)}")
        return []

# Function to get model info from Hugging Face
@st.cache_data(ttl=3600)
def get_model_info(model_id):
    """Get detailed model information from Hugging Face"""
    try:
        url = f"https://huggingface.co/api/models/{model_id}"
        response = requests.get(url, timeout=10)
        
        if response.status_code == 200:
            return response.json()
        return None
    except Exception as e:
        logger.error(f"Error getting model info for {model_id}: {str(e)}")
        return None

# Enhanced error classification with detailed reporting
def classify_error(error_message, model_name):
    """Classify the error and return specific error type and detailed message"""
    error_str = str(error_message).lower()
    original_error = str(error_message)
    
    # API Key errors
    if any(term in error_str for term in ["unauthorized", "invalid token", "api key", "authentication", "401"]):
        return "AUTH_ERROR", f"🔑 **Authentication Error for '{model_name}'**\n\n**Details:** {original_error}\n\n**Solution:** Check your Hugging Face API key is correct and has the right permissions."
    
    # Model not found errors
    if any(term in error_str for term in ["repository not found", "model not found", "does not exist", "404"]):
        return "MODEL_NOT_FOUND", f"🔍 **Model Not Found: '{model_name}'**\n\n**Details:** {original_error}\n\n**Solution:** Verify the model name exists on Hugging Face Hub. Check for typos in model name."
    
    # Task not supported errors
    if any(term in error_str for term in ["task not supported", "not supported for provider", "unsupported"]):
        return "TASK_NOT_SUPPORTED", f"🚫 **Task Not Supported for '{model_name}'**\n\n**Details:** {original_error}\n\n**Solution:** This model doesn't support chat completions or text generation. Try a different model."
    
    # Rate limit errors
    if any(term in error_str for term in ["rate limit", "too many requests", "quota", "429"]):
        return "RATE_LIMIT", f"⏰ **Rate Limit Exceeded for '{model_name}'**\n\n**Details:** {original_error}\n\n**Solution:** Wait a few minutes before trying again."
    
    # Model loading errors
    if any(term in error_str for term in ["model is currently loading", "loading", "initializing", "503"]):
        return "MODEL_LOADING", f"⏳ **Model Loading for '{model_name}'**\n\n**Details:** {original_error}\n\n**Solution:** Wait 30-60 seconds and try again."
    
    # Server errors
    if any(term in error_str for term in ["server error", "internal error", "500", "502", "503"]):
        return "SERVER_ERROR", f"🔧 **Server Error for '{model_name}'**\n\n**Details:** {original_error}\n\n**Solution:** This is a temporary Hugging Face server issue. Try again later."
    
    # Generic model error
    return "MODEL_ERROR", f"❌ **Error for '{model_name}'**\n\n**Details:** {original_error}\n\n**Suggestions:** Try a different model or check the model documentation."

# Simple fallback text generator
def simple_text_completion(query, max_length=100):
    """Simple rule-based text completion as absolute fallback"""
    responses = [
        f"I understand you're asking about: '{query}'. Unfortunately, I cannot provide a proper AI response right now due to model availability issues.",
        f"Your question about '{query}' is noted. I'm currently unable to access AI models, but I acknowledge your inquiry.",
        f"Regarding '{query}' - I'd be happy to help, but all AI models are currently unavailable. Please try again later.",
    ]
    import random
    return random.choice(responses)

# Enhanced Hugging Face integration with automatic fallback models
def query_huggingface_directly(query, model_name, api_key):
    """Query Hugging Face using the chat completions API with intelligent fallbacks"""
    
    if not api_key:
        return {"answer": "🔑 Please provide a valid Hugging Face API key in the sidebar.", "sources": [], "error_type": "AUTH_ERROR"}
    
    try:
        InferenceClient = load_huggingface_client()
        if InferenceClient is None:
            return {"answer": "📦 Hugging Face client not available. Please install huggingface_hub package.", "sources": [], "error_type": "CLIENT_ERROR"}
        
        # Reliable fallback models in order of preference
        fallback_models = [
            "microsoft/DialoGPT-medium",
            "microsoft/DialoGPT-small",
            "facebook/blenderbot-400M-distill",
            "HuggingFaceH4/zephyr-7b-beta",
            "google/flan-t5-base"
        ]
        
        # Start with user's requested model
        models_to_try = [model_name]
        
        # Add fallbacks only if user's model isn't already in the list
        for fallback in fallback_models:
            if fallback != model_name:
                models_to_try.append(fallback)
        
        errors_encountered = []
        
        for i, model in enumerate(models_to_try):
            try:
                client = InferenceClient(api_key=api_key)
                
                # Try chat completions first
                try:
                    completion = client.chat.completions.create(
                        model=model,
                        messages=[
                            {
                                "role": "user",
                                "content": query
                            }
                        ],
                        max_tokens=200,
                        temperature=0.7
                    )
                    
                    # Extract response properly
                    if hasattr(completion, 'choices') and len(completion.choices) > 0:
                        if hasattr(completion.choices[0], 'message'):
                            response_content = completion.choices[0].message.content
                        else:
                            response_content = str(completion.choices[0])
                    else:
                        response_content = str(completion)
                    
                    # Indicate if using fallback
                    if i > 0:
                        response_content = f"🔄 [Using fallback model '{model}' because '{model_name}' failed]: {response_content}"
                    
                    return {"answer": response_content, "sources": [], "error_type": None}
                
                except Exception as chat_error:
                    # Try text generation as fallback
                    try:
                        response = client.text_generation(
                            query,
                            model=model,
                            max_new_tokens=200,
                            temperature=0.7
                        )
                        response_content = str(response).strip()
                        
                        # Indicate if using fallback
                        if i > 0:
                            response_content = f"🔄 [Using fallback model '{model}' because '{model_name}' failed]: {response_content}"
                        
                        return {"answer": response_content, "sources": [], "error_type": None}
                    
                    except Exception as text_error:
                        errors_encountered.append(f"Model {model}: Chat API - {str(chat_error)}, Text API - {str(text_error)}")
                        continue
                    
            except Exception as e:
                error_type, detailed_error_message = classify_error(str(e), model)
                errors_encountered.append(f"Model {model}: {error_type} - {str(e)}")
                
                # For certain critical errors on the first model, don't try fallbacks
                if model == model_name and i == 0:
                    if error_type in ["AUTH_ERROR", "CONNECTION_ERROR"]:
                        return {"answer": detailed_error_message, "sources": [], "error_type": error_type}
                
                continue
        
        # If all models failed
        fallback_response = simple_text_completion(query)
        
        error_summary = f"🚨 **All models failed for user request: '{model_name}'**\n\n"
        error_summary += "**Detailed Error Log:**\n"
        for i, error in enumerate(errors_encountered, 1):
            error_summary += f"{i}. {error}\n"
        
        error_summary += f"\n💡 **Suggestions:**\n"
        error_summary += f"• Search for a different model using the model browser\n"
        error_summary += f"• Check your Hugging Face API key permissions\n"
        error_summary += f"• Wait a few minutes and try again\n"
        error_summary += f"• Try models with 'text-generation' or 'conversational' tags\n\n"
        error_summary += f"**Fallback response:** {fallback_response}"
        
        return {
            "answer": error_summary,
            "sources": [],
            "error_type": "ALL_MODELS_FAILED"
        }
        
    except Exception as e:
        error_type, detailed_error_message = classify_error(str(e), model_name)
        fallback_response = simple_text_completion(query)
        
        full_message = f"{detailed_error_message}\n\n**Fallback response:** {fallback_response}"
        
        return {
            "answer": full_message,
            "sources": [],
            "error_type": error_type
        }

# Helper function for Hugging Face embeddings
def get_huggingface_embedding(text, model_name, api_key):
    """Get embeddings from Hugging Face using feature extraction"""
    try:
        InferenceClient = load_huggingface_client()
        if InferenceClient is None:
            raise Exception("Hugging Face client not available")
        
        client = InferenceClient(api_key=api_key)
        embedding = client.feature_extraction(text, model=model_name)
        
        # Ensure embedding is a flat list
        if isinstance(embedding, list) and len(embedding) > 0:
            if isinstance(embedding[0], list):
                # If it's a nested list, flatten it
                embedding = embedding[0]
        
        return embedding
    except Exception as e:
        logger.error(f"Error getting HF embedding: {str(e)}")
        raise e

# UPDATED Hugging Face document processing function with Excel support
def process_document_huggingface(uploaded_file, table_name="documents", db_path="./lancedb",
                                embedding_model="sentence-transformers/all-MiniLM-L6-v2", 
                                api_key=None, progress_callback=None):
    """Document processing using Hugging Face embeddings with Excel support - FIXED SCHEMA"""
    try:
        if progress_callback:
            progress_callback(0.1, "Extracting text from document...")
            
        file_text = ""
        file_ext = os.path.splitext(uploaded_file.name)[1].lower()
        
        with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp:
            tmp.write(uploaded_file.read())
            tmp_path = tmp.name
            
        try:
            if file_ext == ".txt":
                with open(tmp_path, "r", encoding="utf-8") as f:
                    file_text = f.read()
                    
            elif file_ext == ".pdf":
                try:
                    import PyPDF2
                    with open(tmp_path, "rb") as f:
                        reader = PyPDF2.PdfReader(f)
                        for page in reader.pages:
                            file_text += page.extract_text() + "\n\n"
                except ImportError:
                    file_text = "PyPDF2 not installed. Cannot process PDF files."
                    
            elif file_ext in [".docx", ".doc"]:
                try:
                    import docx
                    doc = docx.Document(tmp_path)
                    paragraphs = []
                    for para in doc.paragraphs:
                        if para.text.strip():
                            paragraphs.append(para.text.strip())
                    file_text = "\n\n".join(paragraphs)
                except ImportError:
                    file_text = "python-docx not installed. Cannot process Word documents."
            
            elif file_ext == ".html":
                try:
                    from bs4 import BeautifulSoup
                    with open(tmp_path, "r", encoding="utf-8") as f:
                        soup = BeautifulSoup(f.read(), 'html.parser')
                        for script in soup(["script", "style"]):
                            script.decompose()
                        file_text = soup.get_text()
                except ImportError:
                    file_text = "BeautifulSoup not installed. Cannot process HTML files."
                    
            elif file_ext in ['.xlsx', '.xls']:
                # NEW: Excel file processing
                file_text = process_excel_file(tmp_path)
                    
            else:
                try:
                    with open(tmp_path, "r", encoding="utf-8") as f:
                        file_text = f.read()
                except:
                    file_text = "Could not extract text from this file type."
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
        
        if not file_text.strip():
            raise ValueError("No text could be extracted from the file")
        
        if progress_callback:
            progress_callback(0.3, "Chunking document...")
            
        # Simple chunking - split by paragraphs with size limit
        chunks = []
        paragraphs = [p.strip() for p in file_text.split("\n\n") if p.strip()]
        
        current_chunk = ""
        chunk_index = 0
        
        for para in paragraphs:
            # If adding this paragraph would make chunk too long, save current chunk
            if len(current_chunk) + len(para) > 1000 and current_chunk:
                chunks.append({
                    "text": current_chunk.strip(),
                    "filename": uploaded_file.name,
                    "title": f"Chunk_{chunk_index + 1}",
                    "source": f"Section_{chunk_index + 1}",
                    "chunk_index": chunk_index
                })
                current_chunk = para
                chunk_index += 1
            else:
                if current_chunk:
                    current_chunk += "\n\n" + para
                else:
                    current_chunk = para
        
        # Add the last chunk
        if current_chunk.strip():
            chunks.append({
                "text": current_chunk.strip(),
                "filename": uploaded_file.name,
                "title": f"Chunk_{chunk_index + 1}",
                "source": f"Section_{chunk_index + 1}",
                "chunk_index": chunk_index
            })
        
        if not chunks:
            raise ValueError("No chunks were created from the document")
        
        if progress_callback:
            progress_callback(0.5, f"Creating embeddings for {len(chunks)} chunks...")
            
        # Generate embeddings using Hugging Face
        InferenceClient = load_huggingface_client()
        if InferenceClient is None:
            raise Exception("Hugging Face client not available")
            
        client = InferenceClient(api_key=api_key)
        processed_chunks = []
        total_chunks = len(chunks)
        
        for i, chunk in enumerate(chunks):
            if i % 5 == 0 and progress_callback:
                progress_callback(0.5 + 0.4 * (i / total_chunks), f"Embedding chunk {i+1}/{total_chunks}")
                
            try:
                embedding = client.feature_extraction(chunk["text"], model=embedding_model)
                
                # Ensure embedding is a flat list
                if isinstance(embedding, list) and len(embedding) > 0:
                    if isinstance(embedding[0], list):
                        # If it's a nested list, flatten it
                        embedding = embedding[0]
                
                # FIXED: Use flat schema - no nested metadata
                processed_chunks.append({
                    "text": chunk["text"],
                    "vector": embedding,
                    "filename": chunk["filename"],
                    "title": chunk["title"],
                    "source": chunk["source"],
                    "chunk_index": chunk["chunk_index"]
                })
                
            except Exception as e:
                logger.error(f"Error creating embedding for chunk {i+1}: {str(e)}")
                # Skip this chunk or use a zero vector
                continue
                
        if not processed_chunks:
            raise ValueError("No embeddings were created successfully")
                
        if progress_callback:
            progress_callback(0.9, "Storing in database...")
            
        # Store in LanceDB with simple schema
        db = lancedb.connect(db_path)
        df = pd.DataFrame(processed_chunks)
        
        if table_name in db.table_names():
            table = db.open_table(table_name)
            table.add(df)
        else:
            table = db.create_table(table_name, df)
            
        if progress_callback:
            progress_callback(1.0, "Processing complete!")
            
        return table, len(chunks)
        
    except Exception as e:
        if progress_callback:
            progress_callback(-1, f"Error: {str(e)}")
        logger.error(f"Error processing document: {str(e)}")
        logger.error(traceback.format_exc())
        raise e

# FIXED Hugging Face RAG implementation with flat schema
def answer_with_rag_huggingface(query, db_path="./lancedb", table_name="documents", 
                               embedding_model="sentence-transformers/all-MiniLM-L6-v2", 
                               llm_model="microsoft/DialoGPT-medium", 
                               top_k=3, api_key=None):
    """Answer a query using RAG with Hugging Face models - FIXED SCHEMA"""
    try:
        # Connect to database
        db = lancedb.connect(db_path)
        
        if table_name not in db.table_names():
            return {"answer": "No documents found in the database. Please process some documents first.", "sources": []}
        
        table = db.open_table(table_name)
        
        # Create query embedding
        query_embedding = get_huggingface_embedding(query, embedding_model, api_key)
        
        # Ensure query embedding is a flat list
        if isinstance(query_embedding, list) and len(query_embedding) > 0:
            if isinstance(query_embedding[0], list):
                query_embedding = query_embedding[0]
        
        results = table.search(query_embedding).limit(top_k).to_pandas()
        
        if len(results) == 0:
            return {"answer": "No relevant documents found.", "sources": []}
        
        # Prepare context from search results
        context_parts = []
        sources = []
        
        for _, row in results.iterrows():
            context_parts.append(row["text"])
            # FIXED: Use flat schema field names
            sources.append({
                "text": row["text"],
                "filename": row.get('filename', 'Unknown'),
                "title": row.get('title', 'Untitled'),
                "source": row.get('source', 'Unknown')
            })
        
        context = "\n\n".join(context_parts)
        
        prompt = f"""Based on the following context, please answer the question. 
        If the answer cannot be found in the context, say so.

Context:
{context}

Question: {query}

Answer:"""

        # Use the improved query function for RAG as well
        result = query_huggingface_directly(prompt, llm_model, api_key)
        result["sources"] = sources
        return result
        
    except Exception as e:
        logger.error(f"Error in RAG pipeline: {str(e)}")
        return {"answer": f"Error generating answer: {str(e)}", "sources": []}

# Function to get current API key based on provider
def get_current_api_key():
    if st.session_state.api_provider == "OpenAI":
        return st.session_state.openai_api_key
    else:
        return st.session_state.hf_api_key

# API Provider and Key input in sidebar
with st.sidebar:
    st.header("API Configuration")
    
    api_provider = st.selectbox(
        "Select API Provider",
        ["Hugging Face", "OpenAI"],
        index=0 if st.session_state.api_provider == "Hugging Face" else 1,
        key="api_provider_select"
    )
    st.session_state.api_provider = api_provider
    
    if api_provider == "OpenAI":
        api_key = st.text_input(
            "Enter your OpenAI API Key", 
            value=st.session_state.openai_api_key,
            type="password",
            help="Get your API key from platform.openai.com/api-keys"
        )
        st.session_state.openai_api_key = api_key
        
        if api_key:
            try:
                OpenAI = load_openai_client()
                if OpenAI:
                    client = OpenAI(api_key=api_key)
                    st.success("✅ Connected to OpenAI API")
                else:
                    st.error("OpenAI client not available")
            except Exception as e:
                st.error(f"Error connecting to OpenAI: {str(e)}")
        else:
            st.warning("⚠️ Please enter your OpenAI API key")
        
        # Set default models for OpenAI
        st.session_state.embedding_model = "text-embedding-3-large"
        st.session_state.llm_model = "gpt-4o-mini"
            
    else:  # Hugging Face
        api_key = st.text_input(
            "Enter your Hugging Face API Key", 
            value=st.session_state.hf_api_key,
            type="password",
            help="Get your API key from huggingface.co/settings/tokens"
        )
        st.session_state.hf_api_key = api_key
        
        if api_key:
            try:
                InferenceClient = load_huggingface_client()
                if InferenceClient:
                    st.success("✅ Connected to Hugging Face API")
                else:
                    st.error("Hugging Face client not available")
            except Exception as e:
                st.error(f"Error connecting to Hugging Face: {str(e)}")
        else:
            st.warning("⚠️ Please enter your Hugging Face API key")
    
    # Model Configuration
    st.header("Model Configuration")
    
    if api_provider == "OpenAI":
        # Load OpenAI functions only when needed
        if st.button("Load OpenAI Models", key="load_openai"):
            with st.spinner("Loading OpenAI functions..."):
                embedding_funcs = load_embedding_functions()
                if embedding_funcs:
                    try:
                        embedding_models = embedding_funcs['get_available_embedding_models']()
                        llm_models = embedding_funcs['get_available_llm_models']()
                        
                        default_embedding = st.selectbox(
                            "Select embedding model", 
                            embedding_models,
                            index=0,
                            key="embedding_model_select"
                        )
                        st.session_state.embedding_model = default_embedding
                        
                        default_llm = st.selectbox(
                            "Select LLM model",
                            llm_models,
                            index=1,
                            key="llm_model_select"
                        )
                        st.session_state.llm_model = default_llm
                    except Exception as e:
                        st.error(f"Error loading OpenAI models: {str(e)}")
                        
        st.info(f"Current embedding model: {st.session_state.embedding_model}")
        st.info(f"Current LLM model: {st.session_state.llm_model}")
        
    else:
        # Hugging Face models
        st.subheader("📝 Embedding Models")
        hf_embedding_models = [
            "sentence-transformers/all-MiniLM-L6-v2",
            "sentence-transformers/all-mpnet-base-v2",
            "sentence-transformers/multi-qa-mpnet-base-dot-v1",
            "sentence-transformers/all-distilroberta-v1"
        ]
        
        default_embedding = st.selectbox(
            "Select embedding model", 
            hf_embedding_models,
            index=0,
            key="hf_embedding_model_select"
        )
        st.session_state.embedding_model = default_embedding
        
        st.subheader("🤖 Language Models")
        
        # Model Search and Browser
        st.markdown("**🔍 Search Any Free HuggingFace Model**")
        
        search_col1, search_col2 = st.columns([3, 1])
        
        with search_col1:
            search_query = st.text_input(
                "Search models",
                placeholder="e.g., 'gpt', 'dialog', 'chat', 'conversation'",
                help="Search for text generation or conversational models"
            )
        
        with search_col2:
            if st.button("🔍 Search", key="search_models"):
                if search_query:
                    with st.spinner("Searching models..."):
                        models = search_huggingface_models(search_query)
                        st.session_state.searched_models = models
                else:
                    st.warning("Enter a search term")
        
        # Display search results
        if st.session_state.searched_models:
            st.markdown(f"**Found {len(st.session_state.searched_models)} models:**")
            
            for model in st.session_state.searched_models[:10]:  # Show top 10
                col1, col2, col3, col4 = st.columns([3, 1, 1, 1])
                
                with col1:
                    st.markdown(f"**{model['id']}**")
                    if model.get('pipeline_tag'):
                        st.caption(f"Type: {model['pipeline_tag']}")
                
                with col2:
                    st.caption(f"⬇️ {model['downloads']:,}")
                
                with col3:
                    st.caption(f"❤️ {model['likes']}")
                
                with col4:
                    if st.button("Use", key=f"use_{model['id']}"):
                        st.session_state.llm_model = model['id']
                        if model['id'] not in st.session_state.favorite_models:
                            st.session_state.favorite_models.append(model['id'])
                        st.success(f"Selected: {model['id']}")
        
        # Current model input and favorites
        st.markdown("**⭐ Quick Select / Custom Input**")
        
        # Show favorite models
        if st.session_state.favorite_models:
            st.markdown("**Your Recent Models:**")
            for fav_model in st.session_state.favorite_models[-5:]:  # Show last 5
                if st.button(f"📌 {fav_model}", key=f"fav_{fav_model}"):
                    st.session_state.llm_model = fav_model
        
        # Manual model input
        manual_model = st.text_input(
            "Or enter model name directly",
            value=st.session_state.llm_model,
            placeholder="e.g., microsoft/DialoGPT-medium",
            help="Enter any Hugging Face model name"
        )
        
        if manual_model != st.session_state.llm_model:
            st.session_state.llm_model = manual_model
            if manual_model and manual_model not in st.session_state.favorite_models:
                st.session_state.favorite_models.append(manual_model)
        
        # Test current model
        current_model = st.session_state.llm_model
        st.markdown(f"**Current Model:** `{current_model}`")
        
        if current_model and st.button("🧪 Test Current Model", key="test_current_model"):
            if st.session_state.hf_api_key:
                with st.spinner("Testing model..."):
                    result = query_huggingface_directly("Hello, how are you?", current_model, st.session_state.hf_api_key)
                    if result.get("error_type"):
                        st.error(f"❌ Test failed: {result['error_type']}")
                        with st.expander("Error Details"):
                            st.markdown(result["answer"])
                    else:
                        st.success("✅ Model works!")
                        st.info(f"Response: {result['answer'][:100]}...")
            else:
                st.warning("Please enter your HuggingFace API key first")
        
        # Popular models quick select
        st.markdown("**🚀 Popular Models (Quick Select)**")
        popular_models = [
            "meta-llama/Llama-3.3-70B-Instruct",
            "zai-org/GLM-4.5",
            "facebook/blenderbot-400M-distill",
            "HuggingFaceH4/zephyr-7b-beta",
            "mistralai/Mistral-7B-Instruct-v0.1",
            "google/flan-t5-base"
        ]
        
        for pop_model in popular_models:
            if st.button(f"⚡ {pop_model}", key=f"pop_{pop_model}"):
                st.session_state.llm_model = pop_model
                if pop_model not in st.session_state.favorite_models:
                    st.session_state.favorite_models.append(pop_model)

# Create tabs for different functionalities
tab1, tab2, tab3 = st.tabs(["Process & Embed", "Chat", "Search"])

with tab1:
    st.header("Process & Create Embeddings")
    st.markdown(f"""
    Upload a file to process, chunk, and create embeddings using **{st.session_state.api_provider}** models.
    
    **📊 Supported file formats:**
    - 📄 **PDF**: `.pdf`
    - 📝 **Text**: `.txt`
    - 📄 **Word**: `.docx`, `.doc`
    - 🌐 **HTML**: `.html`
    - 📊 **Excel**: `.xlsx`, `.xls` ⭐ **NEW!**
    """)
    
    col1, col2 = st.columns(2)
    
    with col1:
        db_path = st.text_input("LanceDB Path", value="./lancedb")
        table_name = st.text_input("Table Name", value="documents")
        
    with col2:
        st.markdown(f"**Embedding Model:** {st.session_state.embedding_model}")
        st.markdown(f"**API Provider:** {st.session_state.api_provider}")
    
    # UPDATED: Added xlsx and xls to supported file types
    uploaded_file_embed = st.file_uploader(
        "Choose a file", 
        type=["pdf", "txt", "docx", "doc", "html", "xlsx", "xls"], 
        key="embed_uploader",
        help="Upload PDF, TXT, DOCX, HTML, or Excel files (.xlsx, .xls)"
    )
    
    if uploaded_file_embed is not None:
        file_details = {
            "Filename": uploaded_file_embed.name,
            "File type": uploaded_file_embed.type,
            "File size": f"{uploaded_file_embed.size:,} bytes"
        }
        st.json(file_details)
        
        if st.button("Process & Create Embeddings", key="embed_file_button"):
            current_api_key = get_current_api_key()
            if not current_api_key:
                st.error(f"Please enter your {st.session_state.api_provider} API key in the sidebar.")
            else:
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                def update_progress(progress, message):
                    if progress < 0:
                        progress_bar.progress(0)
                        status_text.error(message)
                        return
                    progress_bar.progress(progress)
                    status_text.text(message)
                
                try:
                    if st.session_state.api_provider == "OpenAI":
                        with st.spinner("Loading OpenAI functions..."):
                            embedding_funcs = load_embedding_functions()
                        
                        if embedding_funcs:
                            table, num_chunks = embedding_funcs['process_document'](
                                uploaded_file=uploaded_file_embed,
                                table_name=table_name,
                                db_path=db_path,
                                embedding_model=st.session_state.embedding_model,
                                api_key=current_api_key,
                                progress_callback=update_progress
                            )
                        else:
                            st.error("Could not load OpenAI functions")
                            progress_bar.empty()
                            st.stop()
                    else:
                        # Use UPDATED HuggingFace processing function with Excel support
                        table, num_chunks = process_document_huggingface(
                            uploaded_file=uploaded_file_embed,
                            table_name=table_name,
                            db_path=db_path,
                            embedding_model=st.session_state.embedding_model,
                            api_key=current_api_key,
                            progress_callback=update_progress
                        )
                    
                    progress_bar.empty()
                    st.success(f"✅ Document processed successfully! Created {num_chunks} chunks in table '{table_name}'.")
                    
                    # Show file type specific success message
                    file_ext = os.path.splitext(uploaded_file_embed.name)[1].lower()
                    if file_ext in ['.xlsx', '.xls']:
                        st.info("📊 Excel file processed! All sheets and their data have been extracted and embedded.")
                    
                except Exception as e:
                    progress_bar.empty()
                    st.error(f"❌ Error during processing: {str(e)}")

with tab2:
    st.header("Chat")
    
    current_model = st.session_state.llm_model
    st.info(f"🤖 Using: **{current_model}** - Chat Completions API")
    
    chat_mode = st.radio(
        "Select chat mode",
        ["Chat Directly with Model", "Chat with Documents"],
        key="chat_mode"
    )
    
    if "chat_history" not in st.session_state:
        st.session_state.chat_history = []
    
    # Display chat messages
    for message in st.session_state.chat_history:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])
    
    # Chat input
    current_api_key = get_current_api_key()
    if current_api_key:
        user_message = st.chat_input("Ask a question...")
        
        if user_message:
            st.session_state.chat_history.append({"role": "user", "content": user_message})
            
            with st.chat_message("user"):
                st.write(user_message)
            
            with st.spinner("Thinking..."):
                try:
                    if chat_mode == "Chat Directly with Model":
                        if st.session_state.api_provider == "OpenAI":
                            OpenAI = load_openai_client()
                            if OpenAI:
                                client = OpenAI(api_key=current_api_key)
                                response = client.chat.completions.create(
                                    model=st.session_state.llm_model,
                                    messages=[
                                        {"role": "system", "content": "You are a helpful assistant."},
                                        {"role": "user", "content": user_message}
                                    ],
                                    max_tokens=512,
                                    temperature=0.2
                                )
                                result = {"answer": response.choices[0].message.content, "sources": []}
                            else:
                                result = {"answer": "OpenAI client not available", "sources": []}
                        else:
                            result = query_huggingface_directly(
                                query=user_message,
                                model_name=current_model,
                                api_key=current_api_key
                            )
                    else:
                        # RAG Chat
                        if st.session_state.api_provider == "OpenAI":
                            embedding_funcs = load_embedding_functions()
                            if embedding_funcs:
                                db = lancedb.connect("./lancedb")
                                if "documents" not in db.table_names():
                                    result = {"answer": "No documents found. Please process some documents first.", "sources": []}
                                else:
                                    result = embedding_funcs['answer_with_rag'](
                                        query=user_message,
                                        db_path="./lancedb",
                                        table_name="documents",
                                        embedding_model=st.session_state.embedding_model,
                                        llm_model=st.session_state.llm_model,
                                        top_k=3,
                                        api_key=current_api_key
                                    )
                            else:
                                result = {"answer": "OpenAI functions not available", "sources": []}
                        else:
                            db = lancedb.connect("./lancedb")
                            if "documents" not in db.table_names():
                                result = {"answer": "No documents found. Please process some documents first.", "sources": []}
                            else:
                                # Use FIXED HuggingFace RAG function
                                result = answer_with_rag_huggingface(
                                    query=user_message,
                                    db_path="./lancedb",
                                    table_name="documents",
                                    embedding_model=st.session_state.embedding_model,
                                    llm_model=current_model,
                                    top_k=3,
                                    api_key=current_api_key
                                )
                    
                    st.session_state.chat_history.append({"role": "assistant", "content": result["answer"]})
                    
                    with st.chat_message("assistant"):
                        st.markdown(result["answer"])
                    
                    if result.get("error_type"):
                        if result["error_type"] == "ALL_MODELS_FAILED":
                            st.error("🚨 **All Models Failed** - Click below for detailed error analysis")
                            with st.expander("🔍 View Detailed Error Report", expanded=True):
                                st.markdown(result["answer"])
                        else:
                            st.error(f"**Error Type:** {result['error_type']}")
                            with st.expander("🔍 View Error Details", expanded=True):
                                st.markdown(result["answer"])
                    
                    if result.get("sources"):
                        with st.expander("📚 View Sources"):
                            for i, source in enumerate(result["sources"]):
                                st.markdown(f"**Source {i+1}:**")
                                st.markdown(source["text"])
                        
                except Exception as e:
                    error_msg = f"**Unexpected Error:**\n\n```python\n{str(e)}\n```\n\n**Full Traceback:**\n```\n{traceback.format_exc()}\n```"
                    st.error("🔥 **Critical Error Occurred**")
                    
                    with st.expander("🔍 View Full Error Details", expanded=True):
                        st.markdown(error_msg)
                    
                    st.session_state.chat_history.append({"role": "assistant", "content": f"Critical error: {str(e)}"})
                    
                    with st.chat_message("assistant"):
                        st.markdown(f"**Critical Error:** {str(e)}")
    else:
        st.warning(f"Please enter your {st.session_state.api_provider} API key in the sidebar.")
        
    if st.session_state.chat_history:
        if st.button("Clear Chat History", key="clear_chat"):
            st.session_state.chat_history = []
            st.rerun()

with tab3:
    st.header("Semantic Search")
    
    search_query = st.text_input("Enter your search query")
    
    if search_query and st.button("Search", key="search_button"):
        current_api_key = get_current_api_key()
        if not current_api_key:
            st.error(f"Please enter your {st.session_state.api_provider} API key.")
        else:
            with st.spinner("Searching..."):
                try:
                    db = lancedb.connect("./lancedb")
                    
                    if "documents" not in db.table_names():
                        st.error("No documents found. Please process some documents first.")
                    else:
                        table = db.open_table("documents")
                        
                        if st.session_state.api_provider == "OpenAI":
                            embedding_funcs = load_embedding_functions()
                            if embedding_funcs:
                                embedder = embedding_funcs['OpenAIEmbedder'](
                                    model_name=st.session_state.embedding_model, 
                                    api_key=current_api_key
                                )
                                query_embedding = embedder.embed_query(search_query)
                            else:
                                st.error("OpenAI functions not available")
                                st.stop()
                        else:
                            query_embedding = get_huggingface_embedding(
                                search_query,
                                st.session_state.embedding_model,
                                current_api_key
                            )
                        
                        results = table.search(query_embedding).limit(5).to_pandas()
                        
                        st.subheader(f"Found {len(results)} results")
                        
                        for i, row in results.iterrows():
                            with st.expander(f"Result {i+1}: {row.get('filename', 'Unknown')}"):
                                st.markdown(f"**File:** {row.get('filename', 'Unknown')}")
                                st.markdown(f"**Source:** {row.get('source', 'Unknown')}")
                                st.markdown("**Content:**")
                                st.markdown(row["text"])
                                
                                if '_distance' in row.index:
                                    similarity = 1 - float(row['_distance'])
                                    st.progress(similarity)
                                    st.text(f"Similarity: {similarity:.2%}")
                        
                except Exception as e:
                    st.error(f"Error during search: {str(e)}")

# Footer
st.markdown("---")
st.markdown(f"🚀 Powered by {st.session_state.api_provider} and LanceDB | Document Processor with Excel Support 📊")