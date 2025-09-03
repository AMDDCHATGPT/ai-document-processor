import os
import sys
import tempfile
import traceback
import logging
import re
from typing import List, Optional, Dict, Any, Callable
import lancedb
import pandas as pd
from openai import OpenAI

# Set up logging
logging.basicConfig(level=logging.INFO, 
                   format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def get_available_embedding_models():
    """Return list of available OpenAI embedding models"""
    return [
        "text-embedding-3-large",
        "text-embedding-3-small", 
        "text-embedding-ada-002"
    ]

def get_available_llm_models():
    """Return list of available OpenAI language models"""
    return [
        "gpt-4o",
        "gpt-4o-mini",
        "gpt-4-turbo",
        "gpt-4",
        "gpt-3.5-turbo",
        "gpt-3.5-turbo-16k"
    ]

def sanitize_filename(filename: str) -> str:
    """Sanitize filename to be LanceDB compatible"""
    if not filename:
        return "unknown_file"
    
    # Remove path components and get just the filename
    filename = os.path.basename(str(filename))
    
    # Replace problematic characters with safe ones
    sanitized = re.sub(r'[^\w\-\.]', '_', filename)
    sanitized = re.sub(r'_+', '_', sanitized)  # Replace multiple underscores
    sanitized = sanitized.strip('_')  # Remove leading/trailing underscores
    
    # Ensure it's not empty and has reasonable length
    if not sanitized:
        sanitized = "unknown_file"
    
    return sanitized[:100]  # Limit length

def process_excel_file(file_path: str) -> str:
    """Process Excel files and convert to text"""
    try:
        file_ext = os.path.splitext(file_path)[1].lower()
        text_content = []
        
        if file_ext == '.xlsx':
            try:
                import openpyxl
                # Handle .xlsx files with openpyxl
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    text_content.append(f"\n=== Sheet: {sheet_name} ===\n")
                    
                    for row in sheet.iter_rows(values_only=True):
                        row_text = []
                        for cell in row:
                            if cell is not None:
                                row_text.append(str(cell))
                        if row_text:
                            text_content.append(" | ".join(row_text))
            except ImportError:
                return "Error: openpyxl is required to process .xlsx files. Please install it with: pip install openpyxl"
                        
        elif file_ext == '.xls':
            try:
                import xlrd
                # Handle .xls files with xlrd
                workbook = xlrd.open_workbook(file_path)
                for sheet_idx in range(workbook.nsheets):
                    sheet = workbook.sheet_by_index(sheet_idx)
                    text_content.append(f"\n=== Sheet: {sheet.name} ===\n")
                    
                    for row_idx in range(sheet.nrows):
                        row_text = []
                        for col_idx in range(sheet.ncols):
                            cell = sheet.cell(row_idx, col_idx)
                            if cell.value:
                                row_text.append(str(cell.value))
                        if row_text:
                            text_content.append(" | ".join(row_text))
            except ImportError:
                return "Error: xlrd is required to process .xls files. Please install it with: pip install xlrd"
        
        return "\n".join(text_content)
        
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        return f"Error processing Excel file: {str(e)}"

def extract_text_from_file(file_path: str, file_ext: str) -> str:
    """Extract text from various file formats"""
    file_text = ""
    
    if file_ext == ".txt":
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                file_text = f.read()
        except UnicodeDecodeError:
            try:
                with open(file_path, "r", encoding="latin-1") as f:
                    file_text = f.read()
            except Exception as e:
                file_text = f"Error reading text file: {str(e)}"
            
    elif file_ext == ".pdf":
        try:
            import PyPDF2
            with open(file_path, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                for page_num, page in enumerate(reader.pages):
                    page_text = page.extract_text()
                    if page_text.strip():
                        file_text += f"\n\n--- Page {page_num + 1} ---\n\n{page_text}"
        except ImportError:
            file_text = "PyPDF2 not installed. Cannot process PDF files. Install with: pip install PyPDF2"
        except Exception as e:
            file_text = f"Error processing PDF: {str(e)}"
            
    elif file_ext in [".docx", ".doc"]:
        try:
            import docx
            doc = docx.Document(file_path)
            paragraphs = []
            for para in doc.paragraphs:
                if para.text.strip():
                    paragraphs.append(para.text.strip())
            file_text = "\n\n".join(paragraphs)
        except ImportError:
            file_text = "python-docx not installed. Cannot process Word documents. Install with: pip install python-docx"
        except Exception as e:
            file_text = f"Error processing Word document: {str(e)}"
            
    elif file_ext == ".html":
        try:
            from bs4 import BeautifulSoup
            with open(file_path, "r", encoding="utf-8") as f:
                soup = BeautifulSoup(f.read(), 'html.parser')
                # Remove script and style elements
                for script in soup(["script", "style"]):
                    script.decompose()
                file_text = soup.get_text()
        except ImportError:
            file_text = "BeautifulSoup not installed. Cannot process HTML files. Install with: pip install beautifulsoup4"
        except Exception as e:
            file_text = f"Error processing HTML: {str(e)}"
            
    elif file_ext in ['.xlsx', '.xls']:
        file_text = process_excel_file(file_path)
        
    else:
        # Try to read as plain text
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                file_text = f.read()
        except UnicodeDecodeError:
            try:
                with open(file_path, "r", encoding="latin-1") as f:
                    file_text = f.read()
            except Exception as e:
                file_text = f"Could not extract text from this file type: {str(e)}"
    
    return file_text

def chunk_text(text: str, chunk_size: int = 1000, overlap: int = 200) -> List[Dict[str, Any]]:
    """Chunk text into overlapping segments"""
    chunks = []
    
    if not text or not text.strip():
        return chunks
    
    # Split by paragraphs first
    paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
    
    if not paragraphs:
        # Fallback: split by sentences or lines
        lines = [line.strip() for line in text.split("\n") if line.strip()]
        paragraphs = lines if lines else [text]
    
    current_chunk = ""
    chunk_index = 0
    
    for para_index, paragraph in enumerate(paragraphs):
        # If adding this paragraph would exceed chunk size, save current chunk
        if len(current_chunk) + len(paragraph) > chunk_size and current_chunk:
            chunks.append({
                "text": current_chunk.strip(),
                "chunk_index": chunk_index,
                "paragraph_range": f"Para_{max(0, para_index - len(current_chunk.split('\\n\\n')) + 1)}_to_{para_index}"
            })
            
            # Start new chunk with overlap
            words = current_chunk.split()
            if len(words) > overlap // 10:  # Rough word count for overlap
                overlap_text = " ".join(words[-(overlap // 10):])
                current_chunk = overlap_text + "\n\n" + paragraph
            else:
                current_chunk = paragraph
            
            chunk_index += 1
        else:
            if current_chunk:
                current_chunk += "\n\n" + paragraph
            else:
                current_chunk = paragraph
    
    # Add the last chunk
    if current_chunk.strip():
        chunks.append({
            "text": current_chunk.strip(),
            "chunk_index": chunk_index,
            "paragraph_range": f"Para_{max(0, len(paragraphs) - current_chunk.count('\\n\\n'))}_to_{len(paragraphs)}"
        })
    
    # Ensure we have at least one chunk
    if not chunks and text.strip():
        chunks.append({
            "text": text.strip()[:chunk_size],  # Truncate if too long
            "chunk_index": 0,
            "paragraph_range": "Para_1_to_1"
        })
    
    return chunks

class OpenAIEmbedder:
    """OpenAI embedding class with batch processing"""
    
    def __init__(self, model_name: str = "text-embedding-3-large", api_key: str = None):
        self.model_name = model_name
        self.client = OpenAI(api_key=api_key) if api_key else OpenAI()
        
        # Set embedding dimensions based on model
        self.dimensions = {
            "text-embedding-3-large": 3072,
            "text-embedding-3-small": 1536,
            "text-embedding-ada-002": 1536
        }
        self.dimension = self.dimensions.get(model_name, 1536)
        
    def embed_query(self, text: str) -> List[float]:
        """Embed a single query text"""
        try:
            response = self.client.embeddings.create(
                model=self.model_name,
                input=text,
                encoding_format="float"
            )
            return response.data[0].embedding
        except Exception as e:
            logger.error(f"Error creating embedding for query: {str(e)}")
            raise e
    
    def embed_documents(self, texts: List[str], batch_size: int = 100) -> List[List[float]]:
        """Embed multiple documents in batches for efficiency"""
        if not texts:
            return []
            
        embeddings = []
        
        for i in range(0, len(texts), batch_size):
            batch = texts[i:i + batch_size]
            try:
                response = self.client.embeddings.create(
                    model=self.model_name,
                    input=batch,
                    encoding_format="float"
                )
                batch_embeddings = [data.embedding for data in response.data]
                embeddings.extend(batch_embeddings)
                
                logger.info(f"Successfully embedded batch {i//batch_size + 1}/{(len(texts) + batch_size - 1)//batch_size}")
                
            except Exception as e:
                logger.error(f"Error creating embeddings for batch {i//batch_size + 1}: {str(e)}")
                # Fallback to individual embeddings for this batch
                for text in batch:
                    try:
                        single_embedding = self.embed_query(text)
                        embeddings.append(single_embedding)
                    except:
                        # Use zero vector as last resort
                        logger.warning(f"Using zero vector for failed embedding")
                        embeddings.append([0.0] * self.dimension)
        
        return embeddings

def process_document(uploaded_file, 
                    table_name: str = "documents", 
                    db_path: str = "./lancedb",
                    embedding_model: str = "text-embedding-3-large", 
                    api_key: str = None, 
                    progress_callback: Optional[Callable] = None) -> tuple:
    """Process document and create embeddings using OpenAI - FIXED VERSION"""
    
    try:
        if progress_callback:
            progress_callback(0.1, "Extracting text from document...")
            
        # Save uploaded file temporarily
        file_ext = os.path.splitext(uploaded_file.name)[1].lower()
        
        with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp:
            tmp.write(uploaded_file.read())
            tmp_path = tmp.name
            
        try:
            # Extract text based on file type
            file_text = extract_text_from_file(tmp_path, file_ext)
        finally:
            # Always clean up temp file
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
        
        if not file_text or not file_text.strip():
            raise ValueError("No text could be extracted from the file")
            
        if progress_callback:
            progress_callback(0.3, "Chunking document...")
            
        # Chunk the text
        chunks = chunk_text(file_text, chunk_size=1000, overlap=200)
        
        if not chunks:
            raise ValueError("No chunks were created from the document")
            
        if progress_callback:
            progress_callback(0.5, f"Creating embeddings for {len(chunks)} chunks...")
            
        # Initialize embedder
        embedder = OpenAIEmbedder(model_name=embedding_model, api_key=api_key)
        
        # Prepare texts for batch embedding
        texts_to_embed = [chunk["text"] for chunk in chunks]
        
        if progress_callback:
            progress_callback(0.6, "Generating embeddings in batches...")
            
        # Create embeddings in batches (much faster)
        embeddings = embedder.embed_documents(texts_to_embed, batch_size=50)
        
        if len(embeddings) != len(chunks):
            raise ValueError(f"Embedding count mismatch: {len(embeddings)} vs {len(chunks)} chunks")
            
        if progress_callback:
            progress_callback(0.8, "Preparing data for storage...")
            
        # Prepare data for LanceDB using simple format
        processed_chunks = []
        sanitized_filename = sanitize_filename(uploaded_file.name)
        
        for i, (chunk, embedding) in enumerate(zip(chunks, embeddings)):
            processed_chunks.append({
                "text": chunk["text"],
                "vector": embedding,
                "filename": sanitized_filename,
                "title": f"Chunk_{i+1}",
                "source": f"Range_{chunk.get('paragraph_range', f'chunk_{i}')}",
                "chunk_index": chunk.get("chunk_index", i)
            })
            
        if progress_callback:
            progress_callback(0.9, "Storing in database...")
            
        # Store in LanceDB using simple DataFrame approach
        db = lancedb.connect(db_path)
        df = pd.DataFrame(processed_chunks)
        
        if table_name in db.table_names():
            table = db.open_table(table_name)
            table.add(df)
        else:
            table = db.create_table(table_name, df)
            
        if progress_callback:
            progress_callback(1.0, "Processing complete!")
            
        return table, len(chunks)
        
    except Exception as e:
        if progress_callback:
            progress_callback(-1, f"Error: {str(e)}")
        logger.error(f"Error processing document: {str(e)}")
        logger.error(traceback.format_exc())
        raise e

def answer_with_rag(query: str, 
                   db_path: str = "./lancedb", 
                   table_name: str = "documents",
                   embedding_model: str = "text-embedding-3-large", 
                   llm_model: str = "gpt-4o-mini",
                   top_k: int = 3, 
                   api_key: str = None) -> Dict[str, Any]:
    """Answer a query using RAG with OpenAI models - FIXED VERSION"""
    
    try:
        # Connect to database
        db = lancedb.connect(db_path)
        
        if table_name not in db.table_names():
            return {
                "answer": "No documents found in the database. Please process some documents first.",
                "sources": []
            }
        
        table = db.open_table(table_name)
        
        # Create query embedding
        embedder = OpenAIEmbedder(model_name=embedding_model, api_key=api_key)
        query_embedding = embedder.embed_query(query)
        
        # Search for relevant chunks
        results = table.search(query_embedding).limit(top_k).to_pandas()
        
        if len(results) == 0:
            return {
                "answer": "No relevant documents found for your query.",
                "sources": []
            }
        
        # Prepare context and sources
        context_parts = []
        sources = []
        
        for _, row in results.iterrows():
            context_parts.append(row["text"])
            
            sources.append({
                "text": row["text"],
                "filename": row.get("filename", "Unknown"),
                "title": row.get("title", "Untitled"),
                "source": row.get("source", "Unknown")
            })
        
        context = "\n\n".join(context_parts)
        
        # Create prompt for LLM
        system_prompt = """You are a helpful assistant that answers questions based on the provided context.
        Use only the information from the context to answer questions. If you're unsure or the context
        doesn't contain the relevant information, say so clearly.
        
        Provide accurate, helpful answers based on the context provided."""
        
        user_prompt = f"""Based on the following context, please answer the question:

Context:
{context}

Question: {query}

Answer:"""

        # Get response from OpenAI
        client = OpenAI(api_key=api_key) if api_key else OpenAI()
        
        response = client.chat.completions.create(
            model=llm_model,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            max_tokens=1000,
            temperature=0.1
        )
        
        answer = response.choices[0].message.content
        
        return {
            "answer": answer,
            "sources": sources
        }
        
    except Exception as e:
        logger.error(f"Error in RAG pipeline: {str(e)}")
        return {
            "answer": f"Error generating answer: {str(e)}",
            "sources": []
        }

def search_documents(query: str, 
                    db_path: str = "./lancedb", 
                    table_name: str = "documents",
                    embedding_model: str = "text-embedding-3-large",
                    top_k: int = 5,
                    api_key: str = None) -> List[Dict[str, Any]]:
    """Search documents using semantic similarity - FIXED VERSION"""
    
    try:
        # Connect to database
        db = lancedb.connect(db_path)
        
        if table_name not in db.table_names():
            return []
        
        table = db.open_table(table_name)
        
        # Create query embedding
        embedder = OpenAIEmbedder(model_name=embedding_model, api_key=api_key)
        query_embedding = embedder.embed_query(query)
        
        # Search for similar chunks
        results = table.search(query_embedding).limit(top_k).to_pandas()
        
        search_results = []
        for _, row in results.iterrows():
            search_results.append({
                "text": row["text"],
                "filename": row.get("filename", "Unknown"),
                "title": row.get("title", "Untitled"),
                "source": row.get("source", "Unknown"),
                "similarity": 1 - float(row.get("_distance", 1.0))  # Convert distance to similarity
            })
        
        return search_results
        
    except Exception as e:
        logger.error(f"Error searching documents: {str(e)}")
        return []

# Utility functions
def list_available_tables(db_path: str = "./lancedb") -> List[str]:
    """List all available tables in the database"""
    try:
        db = lancedb.connect(db_path)
        return db.table_names()
    except Exception as e:
        logger.error(f"Error listing tables: {str(e)}")
        return []

def get_table_info(table_name: str, db_path: str = "./lancedb") -> Dict[str, Any]:
    """Get information about a specific table"""
    try:
        db = lancedb.connect(db_path)
        if table_name not in db.table_names():
            return {"error": "Table not found"}
        
        table = db.open_table(table_name)
        df = table.to_pandas()
        
        return {
            "num_chunks": len(df),
            "files": df["filename"].unique().tolist() if "filename" in df.columns else [],
            "table_name": table_name
        }
    except Exception as e:
        logger.error(f"Error getting table info: {str(e)}")
        return {"error": str(e)}

def delete_table(table_name: str, db_path: str = "./lancedb") -> bool:
    """Delete a table from the database"""
    try:
        db = lancedb.connect(db_path)
        if table_name in db.table_names():
            db.drop_table(table_name)
            return True
        return False
    except Exception as e:
        logger.error(f"Error deleting table: {str(e)}")
        return False